import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from apps.accounts.utils import get_user_profile, settings_required
from apps.hr_events.forms import HREventForm
from apps.hr_events.hr_events_service import (
    cancel_hr_event,
    create_hr_event,
    get_hr_events_dashboard_data,
)
from apps.hr_events.models import HREvent
from apps.personnel.models import EmployeeProfile


def _visible_employees(profile):
    """Retourne les employes visibles par le RH."""
    employees = EmployeeProfile.objects.select_related("user").exclude(
        user__username="cvbadmin"
    )
    if profile and profile.can_validate_hierarchy:
        employees = employees.filter(department=profile.department)
    return employees


def _employee_search_values(employee):
    return [
        employee.display_name,
        employee.position or "",
        employee.employee_number or "",
        employee.department_name,
    ]


def _normalize_search_text(value):
    import unicodedata

    normalized_value = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(
        character for character in normalized_value if unicodedata.category(character) != "Mn"
    )


def _search_matches(search_term, values):
    normalized_term = _normalize_search_text(search_term)
    if not normalized_term:
        return True
    searchable_text = " ".join(
        _normalize_search_text(value) for value in values if value not in (None, "")
    )
    return normalized_term in searchable_text


def _filter_employees_for_search(employees, search_term):
    if not _normalize_search_text(search_term):
        return employees
    return [emp for emp in employees if _search_matches(search_term, _employee_search_values(emp))]


def _event_table_search_values(event):
    return [
        event.employee.display_name,
        event.employee.employee_number or "",
        event.employee.department_name,
        event.get_event_type_display(),
        event.get_status_display(),
        event.reason or "",
        event.start_date.strftime("%d/%m/%Y") if event.start_date else "",
        event.end_date.strftime("%d/%m/%Y") if event.end_date else "",
        str(event.days),
        event.created_by.username if event.created_by else "",
    ]


def _export_event_rows(events):
    headers = [
        "ID",
        "Employe",
        "Matricule",
        "Type",
        "Statut",
        "Date debut",
        "Date fin",
        "Periode",
        "Nombre de jours",
        "Motif",
        "Cree le",
        "Cree par",
    ]
    rows = []
    for event in events:
        rows.append(
            [
                event.id,
                event.employee.display_name,
                event.employee.employee_number or "",
                event.get_event_type_display(),
                event.get_status_display(),
                event.start_date.strftime("%d/%m/%Y") if event.start_date else "",
                event.end_date.strftime("%d/%m/%Y") if event.end_date else "",
                event.period_label,
                event.display_days,
                event.reason or "",
                event.created_at.strftime("%d/%m/%Y %H:%M") if event.created_at else "",
                event.created_by.username if event.created_by else "-",
            ]
        )
    return headers, rows


def _format_excel_cell(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    return str(value)


def _build_excel_response(filename, sheet_title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    for row in rows:
        sheet.append([_format_excel_cell(value) for value in row])

    header_fill = PatternFill("solid", fgColor="255C3A")
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill("solid", fgColor="F7FBF8")
    clear_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D8E3DB"),
        right=Side(style="thin", color="D8E3DB"),
        top=Side(style="thin", color="D8E3DB"),
        bottom=Side(style="thin", color="D8E3DB"),
    )
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row_fill = zebra_fill if row[0].row % 2 == 0 else clear_fill
        for cell in row:
            cell.alignment = top_alignment
            cell.border = border
            cell.fill = row_fill

    width_map = {}
    for column_index, header in enumerate(headers, start=1):
        candidates = [len(_format_excel_cell(header))]
        for row in rows:
            if column_index - 1 < len(row):
                value = _format_excel_cell(row[column_index - 1])
                candidates.extend(len(part) for part in value.splitlines() or [""])
        width_map[column_index] = min(max(max(candidates) + 3, 12), 42)

    for column_index, width in width_map.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    sheet.row_dimensions[1].height = 28
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 44

    sheet.auto_filter.ref = sheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@settings_required
def hr_events_list_view(request):
    """Liste tous les evenements RH avec filtres et recherche."""
    profile = get_user_profile(request.user)
    search_term = request.GET.get("search", "")
    event_type_filter = request.GET.get("event_type", "")
    status_filter = request.GET.get("status", "")
    employee_filter = request.GET.get("employee", "")

    events = HREvent.objects.select_related(
        "employee", "employee__user", "created_by"
    ).all()

    if profile and profile.can_validate_hierarchy:
        events = events.filter(employee__department=profile.department)

    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    if status_filter:
        events = events.filter(status=status_filter)
    if employee_filter:
        events = events.filter(employee_id=employee_filter)

    # Recherche texte
    if search_term and _normalize_search_text(search_term):
        matching_ids = []
        for event in events:
            values = _event_table_search_values(event)
            if _search_matches(search_term, values):
                matching_ids.append(event.id)
        events = events.filter(id__in=matching_ids)

    events = events.order_by("-created_at")

    # Employes pour le filtre
    employees = _visible_employees(profile)
    employees = _filter_employees_for_search(employees, search_term) if search_term else employees

    context = {
        "events": events,
        "employees": employees,
        "search_term": search_term,
        "event_type_filter": event_type_filter,
        "status_filter": status_filter,
        "employee_filter": employee_filter,
        "event_type_choices": HREvent.TYPE_CHOICES,
        "status_choices": HREvent.STATUS_CHOICES,
        "is_cvbadmin": request.user.username == "cvbadmin",
    }
    return render(request, "hr_events/hr_events_list.html", context)


@login_required
@settings_required
def hr_event_create_view(request):
    """Cree un nouvel evenement RH."""
    profile = get_user_profile(request.user)
    form = HREventForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        employee = form.cleaned_data["employee"]
        event_type = form.cleaned_data["event_type"]
        days = form.cleaned_data["days"]
        start_date = form.cleaned_data.get("start_date")
        end_date = form.cleaned_data.get("end_date")
        start_time = form.cleaned_data.get("start_time")
        end_time = form.cleaned_data.get("end_time")
        reason = form.cleaned_data.get("reason", "")

        success, result = create_hr_event(
            profile=employee,
            event_type=event_type,
            days=days,
            start_date=start_date,
            end_date=end_date,
            start_time=start_time,
            end_time=end_time,
            reason=reason,
            created_by=request.user,
        )

        if success:
            event = result
            type_label = event.get_event_type_display()
            messages.success(
                request,
                f"L'evenement RH ({type_label}) a ete enregistre pour {employee.display_name}.",
            )
            return redirect("hr_events:list")
        else:
            messages.error(request, result)

    context = {
        "form": form,
        "employees": _visible_employees(profile),
    }
    return render(request, "hr_events/hr_event_form.html", context)


@login_required
@settings_required
def hr_event_cancel_view(request, event_id):
    """Annule un evenement RH."""
    if request.method != "POST":
        return redirect("hr_events:list")

    event = get_object_or_404(HREvent, pk=event_id)
    success, message = cancel_hr_event(event)

    if success:
        messages.success(request, message)
    else:
        messages.error(request, message)

    return redirect("hr_events:list")


@login_required
@settings_required
def hr_event_delete_view(request, event_id):
    """Supprime un evenement RH (reserve a cvbadmin)."""
    if request.method != "POST":
        return redirect("hr_events:list")

    if request.user.username != "cvbadmin":
        messages.error(request, "Vous n'avez pas l'autorisation de supprimer cet evenement.")
        return redirect("hr_events:list")

    event = get_object_or_404(HREvent, pk=event_id)
    event.delete()
    messages.success(request, "L'evenement RH a ete supprime definitivement.")
    return redirect("hr_events:list")


@login_required
@settings_required
def hr_event_delete_all_view(request):
    """Supprime tous les evenements RH (reserve a cvbadmin)."""
    if request.method != "POST":
        return redirect("hr_events:list")

    if request.user.username != "cvbadmin":
        messages.error(request, "Vous n'avez pas l'autorisation de supprimer tout l'historique.")
        return redirect("hr_events:list")

    count = HREvent.objects.count()
    HREvent.objects.all().delete()
    messages.success(request, f"Tout l'historique des evenements RH a ete supprime ({count} enregistrement(s)).")
    return redirect("hr_events:list")


@login_required
@settings_required
def hr_event_export_view(request):
    """Exporte les evenements RH en Excel (filtres et recherche appliques)."""
    profile = get_user_profile(request.user)
    search_term = request.GET.get("search", "")
    event_type_filter = request.GET.get("event_type", "")
    status_filter = request.GET.get("status", "")
    employee_filter = request.GET.get("employee", "")

    events = HREvent.objects.select_related(
        "employee", "employee__user", "created_by"
    ).all()

    if profile and profile.can_validate_hierarchy:
        events = events.filter(employee__department=profile.department)

    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    if status_filter:
        events = events.filter(status=status_filter)
    if employee_filter:
        events = events.filter(employee_id=employee_filter)

    # Recherche texte
    if search_term and _normalize_search_text(search_term):
        matching_ids = []
        for event in events:
            values = _event_table_search_values(event)
            if _search_matches(search_term, values):
                matching_ids.append(event.id)
        events = events.filter(id__in=matching_ids)

    events = events.order_by("-created_at")

    headers, rows = _export_event_rows(events)
    return _build_excel_response("evenements-rh.xlsx", "EvenementsRH", headers, rows)