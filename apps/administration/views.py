import io
import json
import unicodedata
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import localtime

from apps.accounts.utils import approval_required, can_manage_settings, settings_required
from apps.administration.forms import (
    ContractTypeForm,
    DepartmentForm,
    EmployeeAccountForm,
    LoginBrandingForm,
    ProjectForm,
    RoleForm,
)
from apps.administration.models import (
    AccountActionHistory,
    LoginBranding,
    RequestActionHistory,
)
from apps.personnel.models import ContractType, Department, EmployeeProfile, Project, Role
from apps.requests_management.models import StaffRequest


def _format_balance_label(value, unit):
    if value is None:
        return f"0 {unit}"
    formatted = format(value, "f").rstrip("0").rstrip(".")
    return f"{formatted or '0'} {unit}"


def _build_balance_distribution(queryset, field_name, unit):
    distribution = (
        queryset.values(field_name)
        .annotate(total=Count("id"))
        .order_by(field_name)
    )
    labels = [_format_balance_label(item[field_name], unit) for item in distribution]
    values = [item["total"] for item in distribution]
    return json.dumps(labels), json.dumps(values)


def _settings_redirect(
    panel="create",
    show_history=False,
    edit_id=None,
    edit_department=None,
    edit_project=None,
    edit_role=None,
    edit_contract_type=None,
):
    url = reverse("administration:settings")
    params = [f"panel={panel}"]
    if show_history:
        params.append("show_history=1")
    if edit_id:
        params.append(f"edit={edit_id}")
    if edit_department:
        params.append(f"edit_department={edit_department}")
    if edit_project:
        params.append(f"edit_project={edit_project}")
    if edit_role:
        params.append(f"edit_role={edit_role}")
    if edit_contract_type:
        params.append(f"edit_contract_type={edit_contract_type}")
    return f"{url}?{'&'.join(params)}"


def _requests_redirect(show_history=False):
    url = reverse("administration:requests")
    if show_history:
        return f"{url}?show_history=1"
    return url


def _visible_employee_queryset(profile):
    employees = EmployeeProfile.objects.select_related("user").exclude(
        user__username="cvbadmin"
    )
    if profile and profile.can_validate_hierarchy:
        employees = employees.filter(department=profile.department)
    return employees


def _scoped_request_queryset(queryset, profile, actionable_only=False):
    if not profile:
        return queryset.none()
    if profile.can_validate_hierarchy:
        queryset = queryset.filter(employee__department=profile.department)
        if actionable_only:
            return queryset.filter(
                status=StaffRequest.STATUS_SUBMITTED,
                approval_stage=StaffRequest.APPROVAL_HIERARCHY,
            )
        return queryset
    if profile.can_manage_settings or profile.can_validate_administration:
        if actionable_only:
            return queryset.filter(
                status=StaffRequest.STATUS_SUBMITTED,
                approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            )
        return queryset
    if profile.can_validate_direction:
        if actionable_only:
            return queryset.filter(
                status=StaffRequest.STATUS_SUBMITTED,
                approval_stage=StaffRequest.APPROVAL_DIRECTION,
            )
        return queryset
    return queryset.none()


def _normalize_search_text(value):
    normalized_value = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(
        character for character in normalized_value if unicodedata.category(character) != "Mn"
    )


def _search_matches(search_term, values):
    normalized_term = _normalize_search_text(search_term)
    if not normalized_term:
        return True
    searchable_text = " ".join(_normalize_search_text(value) for value in values if value not in (None, ""))
    return normalized_term in searchable_text


def _filter_items_for_search(items, search_term, values_getter):
    if not _normalize_search_text(search_term):
        return items
    return [item for item in items if _search_matches(search_term, values_getter(item))]


def _request_period_search_label(item):
    if item.period_label and item.period_label != "-":
        return item.period_label
    if item.start_date:
        return item.start_date.strftime("%d/%m/%Y")
    if item.created_at:
        return localtime(item.created_at).strftime("%d/%m/%Y")
    return ""


def _request_table_search_values(item):
    return [
        item.employee.display_name,
        item.type_label,
        _request_period_search_label(item),
        item.total_days,
        item.reason or item.project_name or "",
        item.get_approval_stage_display(),
        item.admin_comment or "",
        item.status_label,
    ]


def _history_request_table_search_values(item):
    request_item = item["request"]
    stage_statuses = item["stage_statuses"]
    return [
        localtime(request_item.updated_at).strftime("%d/%m/%Y %H:%M") if request_item.updated_at else "",
        request_item.employee.display_name,
        request_item.type_label,
        request_item.total_days,
        _stage_status_label(stage_statuses[0]),
        _stage_status_label(stage_statuses[1]),
        _stage_status_label(stage_statuses[2]),
        request_item.status_label,
        request_item.admin_comment or "",
    ]


def _dashboard_employee_search_values(employee):
    return [
        employee.display_name,
        employee.position or "",
        employee.leave_balance,
        employee.recovery_balance,
    ]


def _account_search_values(employee, current_user):
    values = [
        employee.display_name,
        employee.employee_number or "",
        employee.contract_type_label,
        employee.position or "",
        employee.department_name,
        employee.dashboard_role_label,
        employee.leave_balance,
        employee.recovery_balance,
    ]
    if current_user and employee.user_id == current_user.id:
        values.append("Session en cours")
    return values


def _account_history_search_values(entry):
    return [
        localtime(entry.created_at).strftime("%d/%m/%Y %H:%M") if entry.created_at else "",
        entry.target_username,
        entry.get_action_display(),
        entry.actor.username if entry.actor else "-",
    ]


def _department_search_values(department):
    return [
        department.name,
        department.code or "",
        department.description or "",
        "Actif" if department.is_active else "Inactif",
    ]


def _project_search_values(project):
    return [
        project.name,
        project.code or "",
        project.description or "",
        "Actif" if project.is_active else "Inactif",
    ]


def _role_permissions_search_label(role):
    permissions = []
    if role.can_manage_settings:
        permissions.append("Parametres")
    if role.can_validate_hierarchy:
        permissions.append("Hierarchie")
    if role.can_validate_administration:
        permissions.append("Ressource Humain (RH)")
    if role.can_validate_direction:
        permissions.append("Direction")
    return " ".join(permissions)


def _role_search_values(role):
    return [
        role.label_fr,
        role.code,
        role.get_portal_display(),
        "Visible" if role.show_in_login else "Masque",
        _role_permissions_search_label(role),
        "Actif" if role.is_active else "Inactif",
    ]


def _contract_type_search_values(contract_type):
    return [
        contract_type.label_fr,
        contract_type.code,
        contract_type.order,
        "Actif" if contract_type.is_active else "Inactif",
    ]


def _export_request_rows(requests):
    headers = [
        "ID",
        "Employe",
        "Matricule",
        "Type",
        "Date de creation",
        "Date debut",
        "Date fin",
        "Periode",
        "Nombre de jours",
        "Statut",
        "Etape",
        "Motif / Projet",
        "Commentaire admin",
    ]
    rows = []
    for item in requests:
        created_label = (
            localtime(item.created_at).strftime("%d/%m/%Y %H:%M")
            if item.created_at
            else ""
        )
        rows.append(
            [
                item.id,
                getattr(item.employee, "display_name", ""),
                getattr(item.employee, "employee_number", ""),
                item.type_label,
                created_label,
                item.start_date.strftime("%d/%m/%Y") if item.start_date else "",
                item.end_date.strftime("%d/%m/%Y") if item.end_date else "",
                item.period_label,
                item.total_days or "",
                item.status_label,
                item.get_approval_stage_display(),
                item.reason or item.project_name or "",
                item.admin_comment or "",
            ]
        )
    return headers, rows


def _build_history_requests(current_profile):
    history_requests = (
        StaffRequest.objects.select_related("employee", "employee__user")
        .filter(admin_history__isnull=False)
        .distinct()
        .order_by("-updated_at", "-created_at")
    )
    if current_profile and current_profile.can_validate_hierarchy:
        history_requests = history_requests.filter(employee__department=current_profile.department)
    return [
        {
            "request": history_request,
            "stage_statuses": _build_history_stage_statuses(history_request),
        }
        for history_request in history_requests
    ]


def _can_cancel_history_requests(profile):
    return bool(profile and profile.can_manage_settings)


def _can_delete_history_requests(profile):
    return bool(profile and profile.can_manage_settings)


def _stage_status_label(stage_status):
    username = stage_status.get("username") or "-"
    status = stage_status.get("status") or "-"
    if username == "-":
        return status
    return f"{status} ({username})"


def _export_history_request_rows(history_requests):
    headers = [
        "Date",
        "Employe",
        "Type",
        "Nombre de jours",
        "Chef hierarchique",
        "Ressource Humain (RH)",
        "Direction",
        "Statut final",
        "Commentaire",
    ]
    rows = []
    for item in history_requests:
        request_item = item["request"]
        stage_statuses = item["stage_statuses"]
        rows.append(
            [
                localtime(request_item.updated_at).strftime("%d/%m/%Y %H:%M"),
                request_item.employee.display_name,
                request_item.type_label,
                request_item.total_days or "",
                _stage_status_label(stage_statuses[0]),
                _stage_status_label(stage_statuses[1]),
                _stage_status_label(stage_statuses[2]),
                request_item.status_label,
                request_item.admin_comment or "",
            ]
        )
    return headers, rows


def _export_employee_rows(employees):
    headers = [
        "Employe",
        "Matricule",
        "Poste",
        "Departement",
        "Role",
        "Type de contrat",
        "Conge",
        "Recuperation",
    ]
    rows = []
    for employee in employees:
        rows.append(
            [
                employee.display_name,
                employee.employee_number or "",
                employee.position or "",
                employee.department_name,
                employee.dashboard_role_label,
                employee.contract_type_label,
                employee.leave_balance,
                employee.recovery_balance,
            ]
        )
    return headers, rows


def _export_account_rows(employees):
    headers = [
        "Nom",
        "Nom d'utilisateur",
        "Email",
        "Matricule",
        "Poste",
        "Departement",
        "Role",
        "Type de contrat",
        "Conge",
        "Recuperation",
    ]
    rows = []
    for employee in employees:
        rows.append(
            [
                employee.display_name,
                employee.user.username,
                employee.user.email or "",
                employee.employee_number or "",
                employee.position or "",
                employee.department_name,
                employee.dashboard_role_label,
                employee.contract_type_label,
                employee.leave_balance,
                employee.recovery_balance,
            ]
        )
    return headers, rows


def _export_account_history_rows(entries):
    headers = ["Date", "Compte", "Nom", "Role", "Action", "Utilisateur", "Details"]
    rows = []
    for entry in entries:
        rows.append(
            [
                localtime(entry.created_at).strftime("%d/%m/%Y %H:%M"),
                entry.target_username,
                entry.target_display_name or "",
                entry.target_role or "",
                entry.get_action_display(),
                entry.actor.username if entry.actor else "-",
                entry.details or "",
            ]
        )
    return headers, rows


def _export_department_rows(departments):
    headers = ["Nom", "Code", "Description", "Statut", "Mise a jour"]
    rows = []
    for department in departments:
        rows.append(
            [
                department.name,
                department.code or "",
                department.description or "",
                "Actif" if department.is_active else "Inactif",
                localtime(department.updated_at).strftime("%d/%m/%Y %H:%M"),
            ]
        )
    return headers, rows


def _export_project_rows(projects):
    headers = ["Nom", "Code", "Description", "Statut", "Mise a jour"]
    rows = []
    for project in projects:
        rows.append(
            [
                project.name,
                project.code or "",
                project.description or "",
                "Actif" if project.is_active else "Inactif",
                localtime(project.updated_at).strftime("%d/%m/%Y %H:%M"),
            ]
        )
    return headers, rows


def _export_role_rows(roles):
    headers = [
        "Code",
        "Libelle FR",
        "Portail",
        "Connexion",
        "Parametres",
        "Validation hiérarchie",
        "Validation Ressource Humain (RH)",
        "Validation direction",
        "Statut",
    ]
    rows = []
    for role in roles:
        rows.append(
            [
                role.code,
                role.label_fr,
                role.get_portal_display(),
                "Oui" if role.show_in_login else "Non",
                "Oui" if role.can_manage_settings else "Non",
                "Oui" if role.can_validate_hierarchy else "Non",
                "Oui" if role.can_validate_administration else "Non",
                "Oui" if role.can_validate_direction else "Non",
                "Actif" if role.is_active else "Inactif",
            ]
        )
    return headers, rows


def _export_contract_type_rows(contract_types):
    headers = ["Code", "Libelle FR", "Libelle EN", "Libelle MG", "Statut", "Ordre"]
    rows = []
    for contract_type in contract_types:
        rows.append(
            [
                contract_type.code,
                contract_type.label_fr,
                contract_type.label_en or "",
                contract_type.label_mg or "",
                "Actif" if contract_type.is_active else "Inactif",
                contract_type.order,
            ]
        )
    return headers, rows


def _format_excel_cell(value):
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return format(value, "f").rstrip("0").rstrip(".") or "0"
    if isinstance(value, bool):
        return "Oui" if value else "Non"
    return str(value)


def _build_excel_response(filename, sheet_title, headers, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.append(headers)
    for row in rows:
        sheet.append([_format_excel_cell(value) for value in row])

    header_fill = PatternFill("solid", fgColor="255C3A")
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill("solid", fgColor="F7FBF8")
    clear_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D8E3DB"),
        right=Side(style="thin", color="D8E3DB"),
        top=Side(style="thin", color="D8E3DB"),
        bottom=Side(style="thin", color="D8E3DB"),
    )
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        row_fill = zebra_fill if row[0].row % 2 == 0 else clear_fill
        for cell in row:
            cell.alignment = top_alignment
            cell.border = border
            cell.fill = row_fill

    width_map = {}
    for column_index, header in enumerate(headers, start=1):
        candidates = [len(_format_excel_cell(header))]
        for row in rows:
            if column_index - 1 < len(row):
                value = _format_excel_cell(row[column_index - 1])
                candidates.extend(len(part) for part in value.splitlines() or [""])
        width_map[column_index] = min(max(max(candidates) + 3, 12), 42)

    for column_index, width in width_map.items():
        sheet.column_dimensions[get_column_letter(column_index)].width = width


    sheet.row_dimensions[1].height = 28
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 44

    sheet.auto_filter.ref = sheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _build_admin_dashboard_payload(current_profile):
    employees = _visible_employee_queryset(current_profile)
    actionable_requests = _scoped_request_queryset(
        StaffRequest.objects.all(),
        current_profile,
        actionable_only=True,
    )
    leave_chart_labels, leave_chart_values = _build_balance_distribution(
        employees,
        "leave_balance",
        "jour(s)",
    )
    recovery_chart_labels, recovery_chart_values = _build_balance_distribution(
        employees,
        "recovery_balance",
        "unite(s)",
    )
    return {
        "employees": employees,
        "employee_count": employees.count(),
        "pending_count": actionable_requests.count(),
        "low_leave_count": employees.filter(leave_balance__lt=Decimal("2.0")).count(),
        "low_recovery_count": employees.filter(recovery_balance__lt=Decimal("2.0")).count(),
        "leave_chart_labels": leave_chart_labels,
        "leave_chart_values": leave_chart_values,
        "recovery_chart_labels": recovery_chart_labels,
        "recovery_chart_values": recovery_chart_values,
    }


def _build_requests_overview_context(current_profile, show_history):
    requests = StaffRequest.objects.select_related("employee", "employee__user")
    submitted_requests = _scoped_request_queryset(requests, current_profile, actionable_only=True)
    return {
        "requests": submitted_requests,
        "history_requests": _build_history_requests(current_profile) if show_history else [],
        "can_cancel_history_requests": _can_cancel_history_requests(current_profile),
        "can_delete_history_requests": _can_delete_history_requests(current_profile),
        "pending_count": submitted_requests.count(),
        "show_history": show_history,
    }


def _build_presence_overview_context(current_profile):
    today = timezone.localdate()
    requests = StaffRequest.objects.select_related(
        "employee",
        "employee__user",
        "employee__department",
    ).filter(status=StaffRequest.STATUS_APPROVED)
    requests = _scoped_request_queryset(requests, current_profile, actionable_only=False)

    active_window = Q(start_date__lte=today) & (Q(end_date__gte=today) | Q(end_date__isnull=True))

    leave_requests = (
        requests.filter(request_type=StaffRequest.TYPE_LEAVE)
        .filter(active_window)
        .order_by("start_date", "employee__user__first_name", "employee__user__last_name")
    )
    absence_requests = (
        requests.filter(request_type=StaffRequest.TYPE_ABSENCE)
        .filter(active_window)
        .order_by("start_date", "employee__user__first_name", "employee__user__last_name")
    )

    return {
        "leave_requests": leave_requests,
        "absence_requests": absence_requests,
        "leave_count": leave_requests.count(),
        "absence_count": absence_requests.count(),
        "today": today,
    }


def _build_requests_export_response(requests):
    headers, rows = _export_request_rows(requests)
    return _build_excel_response("demandes.xlsx", "Demandes", headers, rows)


def _build_admin_table_export_response(current_profile, table_key, search_term="", current_user=None):
    is_admin = bool(current_profile and can_manage_settings(current_profile))

    if table_key == "dashboard_employees":
        employees = _visible_employee_queryset(current_profile)
        employees = _filter_items_for_search(employees, search_term, _dashboard_employee_search_values)
        headers, rows = _export_employee_rows(employees)
        return _build_excel_response("soldes-employes.xlsx", "Soldes", headers, rows)

    if table_key == "pending_requests":
        requests = StaffRequest.objects.select_related("employee", "employee__user").prefetch_related("recovery_lines")
        requests = _scoped_request_queryset(requests, current_profile, actionable_only=True)
        requests = requests.order_by("-created_at")
        requests = _filter_items_for_search(requests, search_term, _request_table_search_values)
        headers, rows = _export_request_rows(requests)
        return _build_excel_response("demandes-en-attente.xlsx", "Demandes", headers, rows)

    if table_key == "requests_history":
        history_requests = _build_history_requests(current_profile)
        history_requests = _filter_items_for_search(
            history_requests,
            search_term,
            _history_request_table_search_values,
        )
        headers, rows = _export_history_request_rows(history_requests)
        return _build_excel_response("historique-demandes.xlsx", "Historique", headers, rows)

    if not is_admin:
        return HttpResponse("Non autorise.", status=403)

    if table_key == "accounts":
        employees = (
            EmployeeProfile.objects.select_related("user", "department", "role", "contract_type")
            .exclude(user__username="cvbadmin")
            .order_by("user__first_name", "user__last_name", "user__username")
        )
        employees = _filter_items_for_search(
            employees,
            search_term,
            lambda employee: _account_search_values(employee, current_user),
        )
        headers, rows = _export_account_rows(employees)
        return _build_excel_response("comptes-employes.xlsx", "Comptes", headers, rows)

    if table_key == "accounts_history":
        account_history = AccountActionHistory.objects.select_related("actor")
        account_history = _filter_items_for_search(
            account_history,
            search_term,
            _account_history_search_values,
        )
        headers, rows = _export_account_history_rows(account_history)
        return _build_excel_response("historique-comptes.xlsx", "Historique", headers, rows)

    if table_key == "departments":
        departments = Department.objects.filter(is_active=True).order_by("name")
        departments = _filter_items_for_search(departments, search_term, _department_search_values)
        headers, rows = _export_department_rows(departments)
        return _build_excel_response("departements.xlsx", "Departements", headers, rows)

    if table_key == "projects":
        projects = Project.objects.order_by("name")
        projects = _filter_items_for_search(projects, search_term, _project_search_values)
        headers, rows = _export_project_rows(projects)
        return _build_excel_response("projets.xlsx", "Projets", headers, rows)

    if table_key == "roles":
        roles = Role.objects.order_by("order", "label_fr")
        roles = _filter_items_for_search(roles, search_term, _role_search_values)
        headers, rows = _export_role_rows(roles)
        return _build_excel_response("roles.xlsx", "Roles", headers, rows)

    if table_key == "contract_types":
        contract_types = ContractType.objects.order_by("order", "label_fr")
        contract_types = _filter_items_for_search(
            contract_types,
            search_term,
            _contract_type_search_values,
        )
        headers, rows = _export_contract_type_rows(contract_types)
        return _build_excel_response("types-contrat.xlsx", "TypesContrat", headers, rows)

    return HttpResponse("Export introuvable.", status=404)


def _send_request_email_alert(request_item, branding=None):
    if not branding:
        branding = LoginBranding.objects.first()
    if not branding or not branding.request_submission_email_enabled:
        return False
    recipient = getattr(branding, "email", "").strip()
    if not recipient:
        return False

    subject = "Nouvelle demande de personnel a traiter"
    admin_url = reverse("administration:requests")
    body = (
        f"Une nouvelle demande a ete soumise par {request_item.employee.display_name}.\n"
        f"Type : {request_item.type_label}\n"
        f"Periode : {request_item.start_date or '-'} - {request_item.end_date or '-'}\n"
        f"Nombre de jours : {request_item.total_days}\n"
        f"Motif : {request_item.reason or request_item.project_name or '-'}\n"
        f"Lien : {admin_url}\n"
    )
    try:
        send_mail(subject, body, None, [recipient], fail_silently=True)
        return True
    except Exception:
        return False


def _queue_floating_notification(request, title, message, action_label="", action_url=""):
    request.session["floating_notification"] = {
        "title": title,
        "message": message,
        "action_label": action_label,
        "action_url": action_url,
    }
    request.session.modified = True


def _is_request_in_scope(request_item, profile):
    if not profile:
        return False
    if profile.can_validate_hierarchy:
        return (
            request_item.employee.department_id == profile.department_id
            and request_item.approval_stage == StaffRequest.APPROVAL_HIERARCHY
        )
    if profile.can_manage_settings or profile.can_validate_administration:
        return request_item.approval_stage == StaffRequest.APPROVAL_ADMINISTRATION
    if profile.can_validate_direction:
        return request_item.approval_stage == StaffRequest.APPROVAL_DIRECTION
    return False


def _request_requires_hierarchy(request_item):
    return request_item.employee.role_code == EmployeeProfile.ROLE_USER


def _build_history_stage_statuses(request_item):
    stage_order = [
        StaffRequest.APPROVAL_HIERARCHY,
        StaffRequest.APPROVAL_ADMINISTRATION,
        StaffRequest.APPROVAL_DIRECTION,
    ]
    signature_map = {
        StaffRequest.APPROVAL_HIERARCHY: request_item.hierarchical_signature,
        StaffRequest.APPROVAL_ADMINISTRATION: request_item.administration_signature,
        StaffRequest.APPROVAL_DIRECTION: request_item.direction_signature,
    }
    label_map = {
        StaffRequest.APPROVAL_HIERARCHY: "Chef hierarchique",
        StaffRequest.APPROVAL_ADMINISTRATION: "Ressource Humain (RH)",
        StaffRequest.APPROVAL_DIRECTION: "Direction",
    }
    rejection_stage = request_item.approval_stage if request_item.status == StaffRequest.STATUS_REJECTED else ""
    statuses = []

    for stage_key in stage_order:
        required = (
            stage_key != StaffRequest.APPROVAL_HIERARCHY
            or _request_requires_hierarchy(request_item)
        )
        signature = signature_map[stage_key]

        if not required:
            badge_class = "soft"
            status_label = "Non requise"
        elif rejection_stage == stage_key:
            badge_class = StaffRequest.STATUS_REJECTED
            status_label = "Rejetee"
        elif signature:
            badge_class = StaffRequest.STATUS_APPROVED
            status_label = "Approuvee"
        else:
            badge_class = "soft"
            status_label = "Aucune action"

        statuses.append(
            {
                "label": label_map[stage_key],
                "username": signature or "-",
                "status": status_label,
                "badge_class": badge_class,
            }
        )

    return statuses


def _record_account_history(actor, user, action, details=""):
    profile = getattr(user, "profile", None)
    AccountActionHistory.objects.create(
        actor=actor,
        target_user=user,
        target_username=user.username,
        target_display_name=profile.display_name if profile else "",
        target_role=profile.dashboard_role_label if profile else "",
        action=action,
        details=details,
    )


def _apply_request_balance(request_item):
    profile = request_item.employee
    amount = request_item.total_days or Decimal("0.0")

    if request_item.request_type == StaffRequest.TYPE_LEAVE:
        if profile.leave_balance < amount:
            return (
                False,
                "Solde de conge insuffisant pour approuver cette demande.",
            )
        profile.leave_balance -= amount
        profile.save(update_fields=["leave_balance", "updated_at"])
        return True, "Le solde de conge a ete mis a jour."

    if request_item.request_type == StaffRequest.TYPE_ABSENCE:
        if profile.recovery_balance < amount:
            return (
                False,
                "Solde de recuperation insuffisant pour approuver cette demande d'absence.",
            )
        profile.recovery_balance -= amount
        profile.save(update_fields=["recovery_balance", "updated_at"])
        return True, "Le solde de recuperation a ete mis a jour."

    if request_item.request_type == StaffRequest.TYPE_RECOVERY:
        profile.recovery_balance += amount
        profile.save(update_fields=["recovery_balance", "updated_at"])
        return True, "Le solde de recuperation a ete augmente."

    return True, "La demande a ete approuvee."


def _restore_request_balance_for_cancellation(request_item):
    profile = request_item.employee
    amount = request_item.total_days or Decimal("0.0")

    if request_item.request_type == StaffRequest.TYPE_LEAVE:
        profile.leave_balance += amount
        profile.save(update_fields=["leave_balance", "updated_at"])
        return True, "Le solde de conge a ete restaure."

    if request_item.request_type == StaffRequest.TYPE_ABSENCE:
        profile.recovery_balance += amount
        profile.save(update_fields=["recovery_balance", "updated_at"])
        return True, "Le solde de recuperation a ete restaure."

    if request_item.request_type == StaffRequest.TYPE_RECOVERY:
        if profile.recovery_balance < amount:
            return False, "Impossible d'annuler cette recuperation approuvee car son solde a deja ete utilise."
        profile.recovery_balance -= amount
        profile.save(update_fields=["recovery_balance", "updated_at"])
        return True, "Le solde de recuperation a ete ajuste."

    return True, ""


@login_required
@approval_required
def dashboard_view(request):
    current_profile = getattr(request.user, "profile", None)
    return render(request, "administration/dashboard.html", _build_admin_dashboard_payload(current_profile))


@login_required
@approval_required
def dashboard_data_view(request):
    current_profile = getattr(request.user, "profile", None)
    payload = _build_admin_dashboard_payload(current_profile)
    return JsonResponse(
        {
            "employee_count": payload["employee_count"],
            "pending_count": payload["pending_count"],
            "low_leave_count": payload["low_leave_count"],
            "low_recovery_count": payload["low_recovery_count"],
            "leave_chart_labels": json.loads(payload["leave_chart_labels"]),
            "leave_chart_values": json.loads(payload["leave_chart_values"]),
            "recovery_chart_labels": json.loads(payload["recovery_chart_labels"]),
            "recovery_chart_values": json.loads(payload["recovery_chart_values"]),
            "employees_rows_html": render_to_string(
                "administration/includes/dashboard_employees_rows.html",
                payload,
                request=request,
            ),
        }
    )


@login_required
@approval_required
def requests_overview_view(request):
    current_profile = getattr(request.user, "profile", None)
    show_history = request.GET.get("show_history") == "1"
    return render(
        request,
        "administration/requests_overview.html",
        _build_requests_overview_context(current_profile, show_history),
    )


@login_required
@approval_required
def requests_overview_data_view(request):
    current_profile = getattr(request.user, "profile", None)
    show_history = request.GET.get("show_history") == "1"
    context = _build_requests_overview_context(current_profile, show_history)
    return JsonResponse(
        {
            "pending_count": context["pending_count"],
            "pending_requests_html": render_to_string(
                "administration/includes/pending_requests_rows.html",
                context,
                request=request,
            ),
            "requests_history_html": render_to_string(
                "administration/includes/request_history_rows.html",
                context,
                request=request,
            )
            if show_history
            else "",
        }
    )


@login_required
@approval_required
def presence_overview_view(request):
    current_profile = getattr(request.user, "profile", None)
    return render(
        request,
        "administration/presence_overview.html",
        _build_presence_overview_context(current_profile),
    )


@login_required
@approval_required
def presence_overview_data_view(request):
    current_profile = getattr(request.user, "profile", None)
    context = _build_presence_overview_context(current_profile)
    return JsonResponse(
        {
            "leave_count": context["leave_count"],
            "absence_count": context["absence_count"],
            "leave_rows_html": render_to_string(
                "administration/includes/presence_leave_rows.html",
                context,
                request=request,
            ),
            "absence_rows_html": render_to_string(
                "administration/includes/presence_absence_rows.html",
                context,
                request=request,
            ),
        }
    )


@login_required
@approval_required
def export_requests_view(request, export_format):
    current_profile = getattr(request.user, "profile", None)
    requests = StaffRequest.objects.select_related("employee", "employee__user").prefetch_related("recovery_lines")
    requests = _scoped_request_queryset(requests, current_profile, actionable_only=False)
    requests = requests.order_by("-created_at")
    requests = _filter_items_for_search(requests, request.GET.get("search", ""), _request_table_search_values)
    return _build_requests_export_response(requests)


@login_required
@approval_required
def export_table_view(request, table_key):
    current_profile = getattr(request.user, "profile", None)
    return _build_admin_table_export_response(
        current_profile,
        table_key,
        search_term=request.GET.get("search", ""),
        current_user=request.user,
    )


@login_required
@approval_required
def request_notifications_state_view(request):
    current_profile = getattr(request.user, "profile", None)
    actionable_requests = _scoped_request_queryset(
        StaffRequest.objects.select_related("employee", "employee__user"),
        current_profile,
        actionable_only=True,
    ).order_by("-updated_at", "-created_at")
    pending_count = actionable_requests.count()
    latest_request = actionable_requests.first()
    latest_event_key = ""
    latest_request_payload = None

    if latest_request and latest_request.updated_at:
        latest_event_key = (
            f"{latest_request.id}:{int(localtime(latest_request.updated_at).timestamp())}"
        )
        latest_request_payload = {
            "id": latest_request.id,
            "employee_name": latest_request.employee.display_name,
            "type_label": latest_request.type_label,
            "period_label": latest_request.period_label,
            "updated_at": localtime(latest_request.updated_at).strftime("%d/%m/%Y %H:%M"),
        }

    return JsonResponse(
        {
            "pending_count": pending_count,
            "latest_event_key": latest_event_key,
            "latest_request": latest_request_payload,
        }
    )


@login_required
def acknowledge_request_notification_view(request):
    return JsonResponse({"ok": True})


@login_required
@approval_required
@transaction.atomic
def request_action_view(request, request_id, action):
    if request.method != "POST":
        return redirect("administration:requests")

    request_item = get_object_or_404(
        StaffRequest.objects.select_related("employee", "employee__user"),
        pk=request_id,
    )
    if action != "cancel" and request_item.status != StaffRequest.STATUS_SUBMITTED:
        messages.error(request, "Cette demande a deja ete traitee.")
        return redirect(_requests_redirect(show_history=True))

    comment = request.POST.get("admin_comment", "").strip()
    previous_status = request_item.status
    balance_message = ""

    current_profile = getattr(request.user, "profile", None)
    is_hierarchical = bool(current_profile and current_profile.can_validate_hierarchy)
    is_direction = bool(current_profile and current_profile.can_validate_direction)
    is_admin = bool(current_profile and (current_profile.can_manage_settings or current_profile.can_validate_administration))

    if action == "cancel":
        if not is_admin:
            messages.error(request, "Impossible : seule la Ressource Humain (RH) peut annuler cette demande.")
            return redirect(_requests_redirect(show_history=True))
        if request_item.status != StaffRequest.STATUS_APPROVED:
            messages.error(request, "Cette demande n'est pas approuvee et ne peut pas etre annulee.")
            return redirect(_requests_redirect(show_history=True))
        success, balance_message = _restore_request_balance_for_cancellation(request_item)
        if not success:
            messages.error(request, balance_message)
            return redirect(_requests_redirect(show_history=True))
        request_item.status = StaffRequest.STATUS_CANCELLED
        request_item.approval_stage = StaffRequest.APPROVAL_COMPLETED
        request_item.admin_comment = comment or request_item.admin_comment
        request_item.save(update_fields=[
            "status",
            "approval_stage",
            "admin_comment",
            "updated_at",
        ])
        RequestActionHistory.objects.create(
            request=request_item,
            actor=request.user,
            action=RequestActionHistory.ACTION_REJECTED,
            previous_status=previous_status,
            new_status=request_item.status,
            comment=comment or "Annulation par la Ressource Humain (RH)",
        )
        messages.success(request, "La demande a ete annulee." + (f" {balance_message}" if balance_message else ""))
        return redirect(_requests_redirect(show_history=True))

    if not _is_request_in_scope(request_item, current_profile):
        messages.error(request, "Cette demande n'est pas disponible dans votre perimetre de validation.")
        return redirect(_requests_redirect(show_history=True))

    if action == "reject":
        if not (is_hierarchical or is_admin or is_direction):
            messages.error(request, "Impossible : seuls les validateurs autorises peuvent traiter cette demande.")
            return redirect(_requests_redirect(show_history=True))
        request_item.status = StaffRequest.STATUS_REJECTED
        history_action = RequestActionHistory.ACTION_REJECTED
        confirmation_message = "La demande a ete rejetee."
        if is_hierarchical:
            request_item.hierarchical_signature = request.user.get_username()
        elif is_admin:
            request_item.administration_signature = request.user.get_username()
        elif is_direction:
            request_item.direction_signature = request.user.get_username()
    elif action == "approve":
        if not (is_hierarchical or is_admin or is_direction):
            messages.error(request, "Impossible : seuls les validateurs autorises peuvent traiter cette demande.")
            return redirect("administration:requests")

        if is_hierarchical:
            if request_item.approval_stage != StaffRequest.APPROVAL_HIERARCHY:
                messages.error(
                    request,
                    "Impossible : cette demande n'est pas encore disponible pour le chef hierarchique.",
                )
                return redirect(_requests_redirect(show_history=True))
            request_item.approval_stage = StaffRequest.APPROVAL_ADMINISTRATION
            request_item.hierarchical_signature = request.user.get_username()
            confirmation_message = "La demande a ete transmise a la Ressource Humain (RH)."
            history_action = RequestActionHistory.ACTION_APPROVED
        elif is_admin:
            if request_item.approval_stage == StaffRequest.APPROVAL_HIERARCHY:
                messages.error(
                    request,
                    "Impossible : le chef hierarchique doit d'abord valider cette demande.",
                )
                return redirect(_requests_redirect(show_history=True))
            if request_item.approval_stage == StaffRequest.APPROVAL_DIRECTION:
                messages.error(
                    request,
                    "Impossible : la demande a deja ete transmise a la direction.",
                )
                return redirect(_requests_redirect(show_history=True))
            if request_item.approval_stage != StaffRequest.APPROVAL_ADMINISTRATION:
                messages.error(
                    request,
                    "Impossible : cette demande ne correspond pas a votre etape de validation.",
                )
                return redirect(_requests_redirect(show_history=True))
            request_item.approval_stage = StaffRequest.APPROVAL_DIRECTION
            request_item.administration_signature = request.user.get_username()
            confirmation_message = "La demande a ete transmise a la direction."
            history_action = RequestActionHistory.ACTION_APPROVED
        elif is_direction:
            if request_item.approval_stage == StaffRequest.APPROVAL_HIERARCHY:
                messages.error(
                    request,
                    "Impossible : le chef hierarchique doit d'abord valider cette demande.",
                )
                return redirect(_requests_redirect(show_history=True))
            if request_item.approval_stage == StaffRequest.APPROVAL_ADMINISTRATION:
                messages.error(
                    request,
                    "Impossible : la Ressource Humain (RH) doit d'abord valider cette demande.",
                )
                return redirect(_requests_redirect(show_history=True))
            if request_item.approval_stage != StaffRequest.APPROVAL_DIRECTION:
                messages.error(
                    request,
                    "Impossible : cette demande n'est pas a l'etape de validation de la direction.",
                )
                return redirect(_requests_redirect(show_history=True))
            success, balance_message = _apply_request_balance(request_item)
            if not success:
                messages.error(request, balance_message)
                return redirect("administration:requests")
            request_item.status = StaffRequest.STATUS_APPROVED
            request_item.approval_stage = StaffRequest.APPROVAL_COMPLETED
            request_item.direction_signature = request.user.get_username()
            confirmation_message = "La demande a ete approuvee et finalisee."
            history_action = RequestActionHistory.ACTION_APPROVED
        else:
            messages.error(request, "Impossible : action invalide.")
            return redirect(_requests_redirect(show_history=True))
    else:
        messages.error(request, "Action invalide.")
        return redirect("administration:requests")

    request_item.admin_comment = comment
    request_item.save(update_fields=[
        "status",
        "approval_stage",
        "admin_comment",
        "hierarchical_signature",
        "administration_signature",
        "direction_signature",
        "updated_at",
    ])

    RequestActionHistory.objects.create(
        request=request_item,
        actor=request.user,
        action=history_action,
        previous_status=previous_status,
        new_status=request_item.status,
        comment=comment,
    )

    messages.success(
        request,
        " ".join(
            item
            for item in [confirmation_message, balance_message]
            if item
        ),
    )
    return redirect(_requests_redirect(show_history=True))


@login_required
@settings_required
def request_history_delete_view(request, request_id):
    if request.method != "POST":
        return redirect(_requests_redirect(show_history=True))
    request_item = get_object_or_404(StaffRequest, pk=request_id)
    deleted_count, _ = RequestActionHistory.objects.filter(request=request_item).delete()
    if deleted_count:
        messages.success(request, "L'historique de la demande a ete supprime.")
    else:
        messages.info(request, "Aucune entree d'historique a supprimer pour cette demande.")
    return redirect(_requests_redirect(show_history=True))


@login_required
@settings_required
def account_history_delete_view(request, entry_id):
    if request.method != "POST":
        return redirect(_settings_redirect(panel="accounts", show_history=True))
    if request.user.username != "cvbadmin":
        messages.error(request, "Vous n'avez pas l'autorisation de supprimer cet historique.")
        return redirect(_settings_redirect(panel="accounts", show_history=True))
    entry = get_object_or_404(AccountActionHistory, pk=entry_id)
    entry.delete()
    messages.success(request, "L'entree d'historique a ete supprimee.")
    return redirect(_settings_redirect(panel="accounts", show_history=True))


@login_required
@settings_required
def settings_view(request):
    branding = LoginBranding.objects.first() or LoginBranding.objects.create()
    panel = request.GET.get("panel", "create")
    show_history = request.GET.get("show_history") == "1"
    all_employees = EmployeeProfile.objects.select_related("user", "department", "role", "contract_type")
    employees = all_employees.exclude(user__username="cvbadmin")
    departments = Department.objects.filter(is_active=True).order_by("name")
    projects = Project.objects.order_by("name")
    roles = Role.objects.order_by("order", "label_fr")
    contract_types = ContractType.objects.order_by("order", "label_fr")

    edit_profile = None
    role_editor = None
    contract_type_editor = None
    if request.GET.get("edit"):
        edit_profile = get_object_or_404(all_employees, pk=request.GET["edit"])
        if edit_profile.user.username == "cvbadmin" or edit_profile.user.is_superuser:
            messages.error(request, "Ce compte est protege et ne peut pas etre modifie ici.")
            return redirect(_settings_redirect(panel="accounts", show_history=show_history))
        panel = "accounts"

    department_form = DepartmentForm()
    project_form = ProjectForm()
    role_form = RoleForm()
    contract_type_form = ContractTypeForm()
    account_form = EmployeeAccountForm()
    edit_account_form = EmployeeAccountForm(profile=edit_profile) if edit_profile else None
    branding_form = LoginBrandingForm(instance=branding)

    if request.method == "POST":
        panel = request.POST.get("panel", panel)
        show_history = request.POST.get("show_history") == "1"

        if "create-account" in request.POST:
            account_form = EmployeeAccountForm(request.POST, request.FILES)
            if account_form.is_valid():
                user = account_form.save()
                _record_account_history(
                    request.user,
                    user,
                    AccountActionHistory.ACTION_CREATED,
                    "Compte cree depuis les parametres.",
                )
                messages.success(request, "Le compte employe a ete cree.")
                return redirect(_settings_redirect(panel="create"))

        elif "update-account" in request.POST:
            edit_profile = get_object_or_404(all_employees, pk=request.POST.get("profile_id"))
            if edit_profile.user == request.user:
                messages.error(request, "La modification du compte connecte se fait hors de cet ecran.")
                return redirect(_settings_redirect(panel="accounts"))
            if edit_profile.user.username == "cvbadmin" or edit_profile.user.is_superuser:
                messages.error(request, "Ce compte est protege et ne peut pas etre modifie ici.")
                return redirect(_settings_redirect(panel="accounts", show_history=True))
            edit_account_form = EmployeeAccountForm(
                request.POST,
                request.FILES,
                profile=edit_profile,
            )
            if edit_account_form.is_valid():
                user = edit_account_form.save()
                _record_account_history(
                    request.user,
                    user,
                    AccountActionHistory.ACTION_UPDATED,
                    "Compte modifie depuis les parametres.",
                )
                messages.success(request, "Le compte a ete modifie.")
                return redirect(_settings_redirect(panel="accounts", show_history=True))

        elif "delete-account" in request.POST:
            target_profile = get_object_or_404(all_employees, pk=request.POST.get("profile_id"))
            target_user = target_profile.user
            if target_user == request.user:
                messages.error(request, "Vous ne pouvez pas supprimer votre session en cours.")
                return redirect(_settings_redirect(panel="accounts"))
            if target_user.username == "cvbadmin" or target_user.is_superuser:
                messages.error(request, "Ce compte est protege et ne peut pas etre supprime.")
                return redirect(_settings_redirect(panel="accounts", show_history=True))
            _record_account_history(
                request.user,
                target_user,
                AccountActionHistory.ACTION_DELETED,
                "Compte supprime depuis les parametres.",
            )
            target_user.delete()
            messages.success(request, "Le compte a ete supprime.")
            return redirect(_settings_redirect(panel="accounts", show_history=True))

        elif "save-branding" in request.POST:
            branding_form = LoginBrandingForm(
                request.POST,
                request.FILES,
                instance=branding,
            )
            if branding_form.is_valid():
                branding_form.save()
                messages.success(request, "L'identite visuelle et les preferences d'alerte email ont ete mises a jour.")
                return redirect(_settings_redirect(panel="branding"))

        elif "save-department" in request.POST:
            department_form = DepartmentForm(request.POST)
            if department_form.is_valid():
                department_form.save()
                messages.success(request, "Le departement a ete enregistre.")
                return redirect(_settings_redirect(panel="departments"))

        elif "update-department" in request.POST:
            department = get_object_or_404(Department, pk=request.POST.get("department_id"))
            department_form = DepartmentForm(request.POST, instance=department)
            if department_form.is_valid():
                department_form.save()
                messages.success(request, "Le departement a ete mis a jour.")
                return redirect(_settings_redirect(panel="departments"))

        elif "delete-department" in request.POST:
            department = get_object_or_404(Department, pk=request.POST.get("department_id"))
            department.delete()
            messages.success(request, "Le departement a ete supprime.")
            return redirect(_settings_redirect(panel="departments"))

        elif "save-project" in request.POST:
            project_form = ProjectForm(request.POST)
            if project_form.is_valid():
                project_form.save()
                messages.success(request, "Le projet a ete enregistre.")
                return redirect(_settings_redirect(panel="projects"))

        elif "update-project" in request.POST:
            project = get_object_or_404(Project, pk=request.POST.get("project_id"))
            project_form = ProjectForm(request.POST, instance=project)
            if project_form.is_valid():
                project_form.save()
                messages.success(request, "Le projet a ete mis a jour.")
                return redirect(_settings_redirect(panel="projects"))

        elif "delete-project" in request.POST:
            project = get_object_or_404(Project, pk=request.POST.get("project_id"))
            project.delete()
            messages.success(request, "Le projet a ete supprime.")
            return redirect(_settings_redirect(panel="projects"))

        elif "save-role" in request.POST:
            role_form = RoleForm(request.POST)
            if role_form.is_valid():
                role_form.save()
                messages.success(request, "Le role a ete enregistre.")
                return redirect(_settings_redirect(panel="roles"))

        elif "update-role" in request.POST:
            role_editor = get_object_or_404(Role, pk=request.POST.get("role_id"))
            if role_editor.is_system and role_editor.code == EmployeeProfile.ROLE_ADMIN and not role_editor.can_manage_settings:
                messages.error(request, "Le role Ressource Humain (RH) doit conserver l'acces aux parametres.")
                return redirect(_settings_redirect(panel="roles", edit_role=role_editor.id))
            role_form = RoleForm(request.POST, instance=role_editor)
            if role_form.is_valid():
                updated_role = role_form.save(commit=False)
                if updated_role.is_system and updated_role.code == EmployeeProfile.ROLE_ADMIN:
                    updated_role.can_manage_settings = True
                updated_role.save()
                messages.success(request, "Le role a ete mis a jour.")
                return redirect(_settings_redirect(panel="roles"))

        elif "delete-role" in request.POST:
            role = get_object_or_404(Role, pk=request.POST.get("role_id"))
            if role.is_system:
                messages.error(request, "Ce role systeme ne peut pas etre supprime.")
            elif role.profiles.exists():
                messages.error(request, "Ce role est encore utilise par au moins un compte.")
            else:
                role.delete()
                messages.success(request, "Le role a ete supprime.")
            return redirect(_settings_redirect(panel="roles"))

        elif "save-contract-type" in request.POST:
            contract_type_form = ContractTypeForm(request.POST)
            if contract_type_form.is_valid():
                contract_type_form.save()
                messages.success(request, "Le type de contrat a ete enregistre.")
                return redirect(_settings_redirect(panel="contract_types"))

        elif "update-contract-type" in request.POST:
            contract_type_editor = get_object_or_404(
                ContractType, pk=request.POST.get("contract_type_id")
            )
            contract_type_form = ContractTypeForm(request.POST, instance=contract_type_editor)
            if contract_type_form.is_valid():
                contract_type_form.save()
                messages.success(request, "Le type de contrat a ete mis a jour.")
                return redirect(_settings_redirect(panel="contract_types"))

        elif "delete-contract-type" in request.POST:
            contract_type = get_object_or_404(
                ContractType, pk=request.POST.get("contract_type_id")
            )
            if contract_type.is_system:
                messages.error(request, "Ce type de contrat systeme ne peut pas etre supprime.")
            elif contract_type.profiles.exists():
                messages.error(request, "Ce type de contrat est encore utilise par au moins un compte.")
            else:
                contract_type.delete()
                messages.success(request, "Le type de contrat a ete supprime.")
            return redirect(_settings_redirect(panel="contract_types"))

        if panel == "create" and not account_form.is_bound:
            account_form = EmployeeAccountForm()
        if panel == "branding" and not branding_form.is_bound:
            branding_form = LoginBrandingForm(instance=branding)

    account_history = AccountActionHistory.objects.select_related("actor")
    department_editor = None
    project_editor = None
    if request.GET.get("edit_department"):
        department_editor = get_object_or_404(Department, pk=request.GET["edit_department"])
        department_form = DepartmentForm(instance=department_editor)
    if request.GET.get("edit_project"):
        project_editor = get_object_or_404(Project, pk=request.GET["edit_project"])
        project_form = ProjectForm(instance=project_editor)
    if request.GET.get("edit_role"):
        role_editor = get_object_or_404(Role, pk=request.GET["edit_role"])
        role_form = RoleForm(instance=role_editor)
    if request.GET.get("edit_contract_type"):
        contract_type_editor = get_object_or_404(
            ContractType, pk=request.GET["edit_contract_type"]
        )
        contract_type_form = ContractTypeForm(instance=contract_type_editor)

    return render(
        request,
        "administration/settings.html",
        {
            "account_form": account_form,
            "edit_account_form": edit_account_form,
            "branding_form": branding_form,
            "department_form": department_form,
            "project_form": project_form,
            "role_form": role_form,
            "contract_type_form": contract_type_form,
            "departments": departments,
            "projects": projects,
            "roles": roles,
            "contract_types": contract_types,
            "department_editor": department_editor,
            "project_editor": project_editor,
            "role_editor": role_editor,
            "contract_type_editor": contract_type_editor,
            "employees": employees,
            "account_history": account_history,
            "panel": panel,
            "show_history": show_history,
            "edit_profile": edit_profile,
        },
    )
