import json
from io import BytesIO
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth.models import User
from django.db.models import F
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from apps.accounts.utils import get_role_by_code
from apps.administration.models import AccountActionHistory, LoginBranding, RequestActionHistory
from apps.personnel.leave_service import ensure_leave_window
from apps.personnel.recovery_service import ensure_recovery_window
from apps.personnel.models import Department, EmployeeProfile, Role, AnnualLeave, AnnualRecovery
from apps.requests_management.models import StaffRequest


class AdministrationViewsTests(TestCase):
    def setUp(self):
        admin_role = get_role_by_code(EmployeeProfile.ROLE_ADMIN)
        direction_role = get_role_by_code(EmployeeProfile.ROLE_DIRECTION)
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)

        self.admin = User.objects.create_user(username="admin", password="TestPass123!")
        self.admin.is_staff = True
        self.admin.save()
        self.admin.profile.role = admin_role
        self.admin.profile.recovery_balance = Decimal("5.0")
        self.admin.profile.save()
        self.client.login(username="admin", password="TestPass123!")

        self.employee = User.objects.create_user(
            username="agent",
            password="TestPass123!",
            first_name="Mamy",
            last_name="Agent",
        )
        self.employee.profile.role = user_role
        self.employee.profile.position = "Technicien"
        self.employee.profile.save()
        from apps.personnel.leave_service import ensure_leave_window
        from apps.personnel.recovery_service import ensure_recovery_window
        from apps.personnel.models import AnnualLeave, AnnualRecovery
        ensure_leave_window(self.employee.profile)
        ensure_recovery_window(self.employee.profile)
        AnnualLeave.objects.filter(employee=self.employee.profile).update(consumed=Decimal("0"))
        AnnualRecovery.objects.filter(employee=self.employee.profile).update(balance=Decimal("6.0"), consumed=Decimal("0"))

        self.employee_to_delete = User.objects.create_user(
            username="agent_delete",
            password="TestPass123!",
        )
        self.employee_to_delete.profile.role = user_role
        self.employee_to_delete.profile.save()
        ensure_leave_window(self.employee_to_delete.profile)
        ensure_recovery_window(self.employee_to_delete.profile)
        AnnualLeave.objects.filter(employee=self.employee_to_delete.profile).update(consumed=Decimal("0"))
        AnnualRecovery.objects.filter(employee=self.employee_to_delete.profile).update(balance=Decimal("4.0"), consumed=Decimal("0"))

        self.direction = User.objects.create_user(
            username="direction",
            password="TestPass123!",
        )
        self.direction.profile.role = direction_role
        self.direction.profile.save()

    def _read_workbook_rows(self, response):
        workbook = load_workbook(filename=BytesIO(response.content))
        return list(workbook.active.iter_rows(values_only=True))

    def test_admin_dashboard_requires_admin_role(self):
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)
        user = User.objects.create_user(username="simple", password="TestPass123!")
        user.profile.role = user_role
        user.profile.save()
        self.client.login(username="simple", password="TestPass123!")

        response = self.client.get(reverse("administration:dashboard"), follow=True)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "acces a cette page")

    def test_admin_dashboard_exposes_low_balance_metrics_and_distributions(self):
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)
        low_balance_user = User.objects.create_user(
            username="agent_low",
            password="TestPass123!",
            first_name="Bodo",
            last_name="Petit",
        )
        low_balance_user.profile.role = user_role
        low_balance_user.profile.save()
        ensure_leave_window(low_balance_user.profile)
        ensure_recovery_window(low_balance_user.profile)
        AnnualLeave.objects.filter(employee=low_balance_user.profile).update(consumed=Decimal("28.5"))
        AnnualRecovery.objects.filter(employee=low_balance_user.profile).update(balance=Decimal("1.0"), consumed=Decimal("0"))

        StaffRequest.objects.create(
            employee=low_balance_user.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("1.0"),
            reason="Conge court",
        )

        response = self.client.get(reverse("administration:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["low_leave_count"], 1)
        self.assertEqual(response.context["low_recovery_count"], 1)
        self.assertEqual(response.context["pending_count"], 1)
        self.assertIn("1.5 jour(s)", json.loads(response.context["leave_chart_labels"]))
        self.assertIn("1 unite(s)", json.loads(response.context["recovery_chart_labels"]))

    def test_admin_dashboard_hides_cvbadmin_from_stats_and_tables(self):
        admin_role = get_role_by_code(EmployeeProfile.ROLE_ADMIN)
        hidden_admin = User.objects.create_user(
            username="cvbadmin",
            password="TestPass123!",
            first_name="Compte",
            last_name="Cache",
        )
        hidden_admin.profile.role = admin_role
        hidden_admin.profile.save()
        ensure_leave_window(hidden_admin.profile)
        ensure_recovery_window(hidden_admin.profile)
        AnnualLeave.objects.filter(employee=hidden_admin.profile).update(consumed=Decimal("30.0"))
        AnnualRecovery.objects.filter(employee=hidden_admin.profile).update(balance=Decimal("0.0"), consumed=Decimal("0"))

        response = self.client.get(reverse("administration:dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["employee_count"], 3)
        self.assertNotContains(response, "Compte Cache")

    def test_admin_transmits_leave_request_to_direction(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("2.0"),
            reason="Conge familial",
        )

        response = self.client.post(
            reverse("administration:request_action", args=[staff_request.id, "approve"]),
            {"admin_comment": "Validation admin accordee."},
            follow=True,
        )

        staff_request.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(staff_request.status, StaffRequest.STATUS_SUBMITTED)
        self.assertEqual(staff_request.approval_stage, StaffRequest.APPROVAL_DIRECTION)
        self.assertEqual(self.employee.profile.leave_balance, Decimal("60.0"))
        self.assertTrue(
            RequestActionHistory.objects.filter(
                request=staff_request,
                action=RequestActionHistory.ACTION_APPROVED,
                actor=self.admin,
            ).exists()
        )

    def test_admin_transmits_absence_request_to_direction(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_ABSENCE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("2.0"),
            reason="Absence terrain",
        )
        ensure_recovery_window(self.employee.profile)
        AnnualRecovery.objects.filter(employee=self.employee.profile).update(balance=Decimal("2.0"))

        response = self.client.post(
            reverse("administration:request_action", args=[staff_request.id, "approve"]),
            {"admin_comment": "Absence transmise a la direction."},
            follow=True,
        )

        staff_request.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(staff_request.status, StaffRequest.STATUS_SUBMITTED)
        self.assertEqual(staff_request.approval_stage, StaffRequest.APPROVAL_DIRECTION)
        from apps.personnel.recovery_service import get_recovery_balance
        self.assertEqual(get_recovery_balance(self.employee.profile), Decimal("6.0"))

    def test_admin_transmits_recovery_request_to_direction(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_RECOVERY,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("3.0"),
            project_name="Mission",
        )
        ensure_recovery_window(self.employee.profile)
        AnnualRecovery.objects.filter(employee=self.employee.profile).update(balance=Decimal("3.0"))

        response = self.client.post(
            reverse("administration:request_action", args=[staff_request.id, "approve"]),
            {"admin_comment": "Recuperation transmise a la direction."},
            follow=True,
        )

        staff_request.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(staff_request.status, StaffRequest.STATUS_SUBMITTED)
        self.assertEqual(staff_request.approval_stage, StaffRequest.APPROVAL_DIRECTION)
        from apps.personnel.recovery_service import get_recovery_balance
        self.assertEqual(get_recovery_balance(self.employee.profile), Decimal("9.0"))

    def test_admin_can_cancel_approved_request_and_restore_balance(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_APPROVED,
            approval_stage=StaffRequest.APPROVAL_COMPLETED,
            total_days=Decimal("2.0"),
            reason="Conge annuel",
            admin_comment="Deja approuvee.",
        )
        ensure_leave_window(self.employee.profile)
        AnnualLeave.objects.filter(employee=self.employee.profile).update(consumed=Decimal("2.0"))

        response = self.client.post(
            reverse("administration:request_action", args=[staff_request.id, "cancel"]),
            follow=True,
        )

        staff_request.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(staff_request.status, StaffRequest.STATUS_CANCELLED)
        from apps.personnel.leave_service import get_leave_balance
        self.assertEqual(get_leave_balance(self.employee.profile), Decimal("56.0"))

    def test_requests_history_shows_cancel_action_for_admin_in_page_and_ajax_rows(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_APPROVED,
            approval_stage=StaffRequest.APPROVAL_COMPLETED,
            total_days=Decimal("2.0"),
            reason="Conge annuel",
        )
        RequestActionHistory.objects.create(
            request=staff_request,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_APPROVED,
            comment="Validation finale.",
        )

        overview_response = self.client.get(
            f"{reverse('administration:requests')}?show_history=1"
        )
        ajax_response = self.client.get(
            f"{reverse('administration:requests_overview_data')}?show_history=1",
            HTTP_X_REQUESTED_WITH="fetch",
        )

        cancel_url = reverse("administration:request_action", args=[staff_request.id, "cancel"])

        self.assertEqual(overview_response.status_code, 200)
        self.assertContains(overview_response, "Annuler")
        self.assertContains(overview_response, cancel_url)
        self.assertEqual(ajax_response.status_code, 200)
        self.assertIn(cancel_url, ajax_response.json()["requests_history_html"])

    def test_requests_history_hides_cancel_action_for_direction(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_APPROVED,
            approval_stage=StaffRequest.APPROVAL_COMPLETED,
            total_days=Decimal("2.0"),
            reason="Conge annuel",
            direction_signature="direction",
        )
        RequestActionHistory.objects.create(
            request=staff_request,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_APPROVED,
            comment="Validation finale.",
        )

        self.client.logout()
        self.client.login(username="direction", password="TestPass123!")

        overview_response = self.client.get(
            f"{reverse('administration:requests')}?show_history=1"
        )
        ajax_response = self.client.get(
            f"{reverse('administration:requests_overview_data')}?show_history=1",
            HTTP_X_REQUESTED_WITH="fetch",
        )

        cancel_url = reverse("administration:request_action", args=[staff_request.id, "cancel"])

        self.assertEqual(overview_response.status_code, 200)
        self.assertNotContains(overview_response, "<th>Annuler</th>", html=True)
        self.assertNotContains(overview_response, cancel_url)
        self.assertEqual(ajax_response.status_code, 200)
        self.assertNotIn(cancel_url, ajax_response.json()["requests_history_html"])

    def test_direction_cannot_cancel_approved_request(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_APPROVED,
            approval_stage=StaffRequest.APPROVAL_COMPLETED,
            total_days=Decimal("2.0"),
            reason="Conge annuel",
        )
        starting_balance = self.employee.profile.leave_balance

        self.client.logout()
        self.client.login(username="direction", password="TestPass123!")

        response = self.client.post(
            reverse("administration:request_action", args=[staff_request.id, "cancel"]),
            follow=True,
        )

        staff_request.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(staff_request.status, StaffRequest.STATUS_APPROVED)
        self.assertEqual(self.employee.profile.leave_balance, starting_balance)
        self.assertContains(response, "seule la Ressource Humain (RH) peut annuler cette demande")

    def test_request_history_groups_multiple_actions_on_single_row(self):
        staff_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_DIRECTION,
            total_days=Decimal("2.0"),
            reason="Conge annuel",
            admin_comment="En attente de la direction.",
            hierarchical_signature="chef-service",
            administration_signature="admin",
        )
        RequestActionHistory.objects.create(
            request=staff_request,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_SUBMITTED,
            comment="Validation admin.",
        )
        RequestActionHistory.objects.create(
            request=staff_request,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_SUBMITTED,
            comment="Validation chef deja enregistree.",
        )

        response = self.client.get(
            f"{reverse('administration:requests')}?show_history=1"
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["history_requests"]), 1)
        history_row = response.context["history_requests"][0]
        self.assertEqual(history_row["request"].id, staff_request.id)
        self.assertEqual(history_row["stage_statuses"][0]["status"], "Approuvee")
        self.assertEqual(history_row["stage_statuses"][1]["status"], "Approuvee")
        self.assertEqual(history_row["stage_statuses"][2]["status"], "Aucune action")

    def test_requests_overview_displays_total_days_in_pending_and_history_tables(self):
        pending_request = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_ABSENCE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("4.0"),
            reason="Mission terrain",
            administration_signature="admin",
        )
        RequestActionHistory.objects.create(
            request=pending_request,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_SUBMITTED,
            comment="Transmission en cours.",
        )

        response = self.client.get(
            f"{reverse('administration:requests')}?show_history=1"
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Nombre de jours")
        self.assertContains(response, "4,0")

    def test_admin_can_create_account_with_contract_type(self):
        response = self.client.post(
            reverse("administration:settings"),
            {
                "panel": "create",
                "create-account": "1",
                "username": "nouvel-agent",
                "password": "TestPass123!",
                "first_name": "Jean",
                "last_name": "Rakoto",
                "email": "jean@example.com",
                "employee_number": "EMP-010",
                "position": "Technicien",
                "contract_type": EmployeeProfile.CONTRACT_TYPE_CDI,
                "leave_balance": "15.0",
                "recovery_balance": "2.0",
                "role": EmployeeProfile.ROLE_USER,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        created_user = User.objects.get(username="nouvel-agent")
        self.assertEqual(created_user.profile.contract_type, EmployeeProfile.CONTRACT_TYPE_CDI)

    def test_admin_can_update_existing_account_and_log_history(self):
        response = self.client.post(
            reverse("administration:settings"),
            {
                "panel": "accounts",
                "profile_id": self.employee.profile.id,
                "update-account": "1",
                "username": "agent",
                "password": "",
                "first_name": "Mamy",
                "last_name": "Agent Modifie",
                "email": "agent@example.com",
                "employee_number": "EMP-004",
                "position": "Responsable terrain",
                "contract_end_date": "",
                "leave_balance": "12.0",
                "recovery_balance": "5.0",
                "role": EmployeeProfile.ROLE_USER,
            },
            follow=True,
        )

        self.employee.refresh_from_db()
        self.employee.profile.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.employee.last_name, "Agent Modifie")
        self.assertEqual(self.employee.profile.position, "Responsable terrain")
        self.assertTrue(
            AccountActionHistory.objects.filter(
                target_username="agent",
                action=AccountActionHistory.ACTION_UPDATED,
                actor=self.admin,
            ).exists()
        )

    def test_admin_can_delete_account_and_log_history(self):
        response = self.client.post(
            reverse("administration:settings"),
            {
                "panel": "accounts",
                "show_history": "1",
                "profile_id": self.employee_to_delete.profile.id,
                "delete-account": "1",
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(User.objects.filter(username="agent_delete").exists())
        self.assertTrue(
            AccountActionHistory.objects.filter(
                target_username="agent_delete",
                action=AccountActionHistory.ACTION_DELETED,
                actor=self.admin,
            ).exists()
        )

    def test_export_requests_returns_excel_file_even_for_csv_route(self):
        StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("2.0"),
            reason="Conge test",
        )

        response = self.client.get(
            reverse("administration:export_requests", args=["csv"])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("demandes.xlsx", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

    def test_export_table_accounts_returns_excel_file(self):
        response = self.client.get(
            reverse("administration:export_table", args=["accounts"])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("comptes-employes.xlsx", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

    def test_export_table_accounts_applies_search_from_visible_table(self):
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)
        searched_user = User.objects.create_user(
            username="tendry",
            password="TestPass123!",
            first_name="Tendry",
            last_name="Rakoto",
        )
        searched_user.profile.role = user_role
        searched_user.profile.position = "Analyste"
        searched_user.profile.leave_balance = Decimal("8.0")
        searched_user.profile.recovery_balance = Decimal("2.0")
        searched_user.profile.save()

        response = self.client.get(
            reverse("administration:export_table", args=["accounts"]),
            {"search": "tendry"},
        )

        self.assertEqual(response.status_code, 200)
        rows = self._read_workbook_rows(response)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], searched_user.profile.display_name)
        self.assertEqual(rows[1][4], "Analyste")

    def test_export_table_departments_matches_displayed_active_rows(self):
        Department.objects.create(name="Administration", code="ADM", is_active=True)
        Department.objects.create(name="Archive", code="ARC", is_active=False)

        response = self.client.get(
            reverse("administration:export_table", args=["departments"])
        )

        self.assertEqual(response.status_code, 200)
        rows = self._read_workbook_rows(response)
        exported_names = [row[0] for row in rows[1:]]
        self.assertIn("Administration", exported_names)
        self.assertNotIn("Archive", exported_names)

    def test_export_table_requests_history_returns_excel_file(self):
        request_item = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_ABSENCE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("1.0"),
            reason="Mission",
            hierarchical_signature="chef",
            administration_signature="admin",
        )
        RequestActionHistory.objects.create(
            request=request_item,
            actor=self.admin,
            action=RequestActionHistory.ACTION_APPROVED,
            previous_status=StaffRequest.STATUS_SUBMITTED,
            new_status=StaffRequest.STATUS_SUBMITTED,
            comment="Validation",
        )

        response = self.client.get(
            reverse("administration:export_table", args=["requests_history"])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("historique-demandes.xlsx", response["Content-Disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

    def test_branding_settings_displays_email_alert_toggle(self):
        response = self.client.get(
            f"{reverse('administration:settings')}?panel=branding"
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Alertes email a la soumission")
        self.assertContains(response, "request_submission_email_enabled")

    def test_request_notifications_state_returns_pending_request_summary(self):
        request_item = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_ABSENCE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("1.0"),
            reason="Mission",
        )

        response = self.client.get(reverse("administration:request_notifications_state"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["pending_count"], 1)
        self.assertIn(str(request_item.id), payload["latest_event_key"])
        self.assertEqual(payload["latest_request"]["employee_name"], "Mamy Agent")

    def test_request_email_alert_respects_admin_toggle(self):
        branding = LoginBranding.objects.create(
            site_name="Centre ValBio",
            subtitle="Gestion",
            email="admin@example.com",
            request_submission_email_enabled=False,
        )
        request_item = StaffRequest.objects.create(
            employee=self.employee.profile,
            request_type=StaffRequest.TYPE_LEAVE,
            status=StaffRequest.STATUS_SUBMITTED,
            approval_stage=StaffRequest.APPROVAL_ADMINISTRATION,
            total_days=Decimal("1.0"),
            reason="Conge court",
        )

        with patch("apps.administration.views.send_mail") as mocked_send_mail:
            from apps.administration.views import _send_request_email_alert

            result = _send_request_email_alert(request_item, branding=branding)

        self.assertFalse(result)
        mocked_send_mail.assert_not_called()

    def test_calendar_view_allows_employee_selection_for_admin(self):
        admin_role, _ = Role.objects.get_or_create(
            code=EmployeeProfile.ROLE_ADMIN,
            defaults={
                "label_fr": "Administrateur",
                "portal": Role.PORTAL_ADMIN,
                "can_manage_settings": True,
            },
        )
        admin_user = User.objects.create_user(username="calendar_admin", password="TestPass123!")
        admin_user.profile.role = admin_role
        admin_user.profile.save()

        employee_role, _ = Role.objects.get_or_create(
            code=EmployeeProfile.ROLE_USER,
            defaults={
                "label_fr": "Employe",
                "portal": Role.PORTAL_EMPLOYEE,
            },
        )
        employee_user = User.objects.create_user(username="calendar_emp", password="TestPass123!")
        employee_user.profile.role = employee_role
        employee_user.profile.save()

        self.client.logout()
        self.client.login(username="calendar_admin", password="TestPass123!")

        response = self.client.get(
            reverse("administration:calendar") + f"?year=2026&employee={employee_user.profile.id}"
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, employee_user.profile.display_name)
        self.assertContains(response, "calendarEmployeeSelect")

    def test_calendar_employee_search_returns_visible_employees(self):
        admin_role, _ = Role.objects.get_or_create(
            code=EmployeeProfile.ROLE_ADMIN,
            defaults={
                "label_fr": "Administrateur",
                "portal": Role.PORTAL_ADMIN,
                "can_manage_settings": True,
            },
        )
        admin_user = User.objects.create_user(username="calendar_admin2", password="TestPass123!")
        admin_user.profile.role = admin_role
        admin_user.profile.save()

        employee_role, _ = Role.objects.get_or_create(
            code=EmployeeProfile.ROLE_USER,
            defaults={
                "label_fr": "Employe",
                "portal": Role.PORTAL_EMPLOYEE,
            },
        )
        employee_user = User.objects.create_user(username="calendar_emp2", password="TestPass123!")
        employee_user.profile.role = employee_role
        employee_user.profile.save()

        self.client.logout()
        self.client.login(username="calendar_admin2", password="TestPass123!")

        response = self.client.get(
            reverse("administration:calendar_employee_search") + "?term=" + employee_user.profile.display_name[:5]
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("results", data)
        matched_ids = [item["id"] for item in data["results"]]
        self.assertIn(employee_user.profile.id, matched_ids)


class ExceptionalAbsencesManagementTests(TestCase):
    """Tests for the RH exceptional absences management interface."""

    def setUp(self):
        admin_role = get_role_by_code(EmployeeProfile.ROLE_ADMIN)
        direction_role = get_role_by_code(EmployeeProfile.ROLE_DIRECTION)
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)

        self.admin = User.objects.create_user(username="rh", password="TestPass123!")
        self.admin.is_staff = True
        self.admin.save()
        self.admin.profile.role = admin_role
        self.admin.profile.save()
        self.client.login(username="rh", password="TestPass123!")

        self.direction = User.objects.create_user(
            username="direction", password="TestPass123!", first_name="Dir", last_name="Ector"
        )
        self.direction.profile.role = direction_role
        self.direction.profile.save()

        self.employee = User.objects.create_user(
            username="agent", password="TestPass123!", first_name="Mamy", last_name="Agent"
        )
        self.employee.profile.role = user_role
        self.employee.profile.employee_number = "EMP001"
        self.employee.profile.save()
        ensure_leave_window(self.employee.profile)
        ensure_recovery_window(self.employee.profile)
        AnnualLeave.objects.filter(employee=self.employee.profile).update(consumed=Decimal("0"))

        self.employee_no_access = User.objects.create_user(
            username="regular_agent", password="TestPass123!", first_name="John", last_name="Doe"
        )
        self.employee_no_access.profile.role = user_role
        self.employee_no_access.profile.save()

    def _create_exceptional_request(self, days="3", start="2026-07-13", end="2026-07-15", acknowledged=True):
        self.client.logout()
        self.client.login(username="agent", password="TestPass123!")
        ensure_leave_window(self.employee.profile)
        AnnualLeave.objects.filter(employee=self.employee.profile).update(
            consumed=F("quota")
        )
        post_data = {
            "start_date": start,
            "end_date": end,
            "total_days": days,
            "remaining_days_for_reason": "",
            "reason": "Conge exceptionnel",
        }
        if acknowledged:
            post_data["acknowledged_salary_deduction"] = "on"
        self.client.post(reverse("requests_management:leave_create"), post_data, follow=True)
        req = StaffRequest.objects.first()
        if req and req.approval_stage != StaffRequest.APPROVAL_ADMINISTRATION:
            req.approval_stage = StaffRequest.APPROVAL_ADMINISTRATION
            req.save(update_fields=["approval_stage"])
        return req

    def _get_admin(self):
        self.client.logout()
        self.client.login(username="rh", password="TestPass123!")

    def test_exceptional_absences_page_accessible_to_admin(self):
        self._create_exceptional_request()
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)

    def test_unauthenticated_redirects_to_login(self):
        self.client.logout()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 302)

    def test_non_admin_cannot_access_page(self):
        self._get_admin()
        self.client.logout()
        self.client.login(username="regular_agent", password="TestPass123!")
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 302)

    def test_created_exceptional_request_appears_in_table(self):
        req = self._create_exceptional_request(days="3", start="2026-07-13", end="2026-07-15")
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mamy Agent")
        self.assertContains(response, "13/07/2026")
        self.assertContains(response, "3,0")

    def test_search_by_employee_name(self):
        self._create_exceptional_request()
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences") + "?q=Mamy")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mamy")

    def test_search_by_employee_number(self):
        self._create_exceptional_request()
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences") + "?q=EMP001")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agent")

    def test_filter_by_status(self):
        self._create_exceptional_request()
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences") + "?status=submitted")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agent")

    def test_filter_by_year(self):
        self._create_exceptional_request()
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences") + "?year=2026")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agent")

    def test_pagination_with_multiple_requests(self):
        for _ in range(25):
            self._create_exceptional_request(days="3", start="2026-07-13", end="2026-07-15")
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Precedent")
        self.assertContains(response, "Suivant")

    def test_empty_state(self):
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aucune absence exceptionnelle trouvee.")

    def test_detail_view_shows_exceptional_info(self):
        req = self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absence_detail", args=[req.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Absence exceptionnelle")
        self.assertContains(response, "Retenue salariale")
        self.assertContains(response, "3,0")
        self.assertContains(response, "Oui")

    def test_detail_view_permission_denied_for_employee(self):
        req = self._create_exceptional_request(days="3")
        self.client.logout()
        self.client.login(username="regular_agent", password="TestPass123!")
        response = self.client.get(reverse("administration:exceptional_absence_detail", args=[req.id]))
        self.assertEqual(response.status_code, 302)

    def test_approve_uses_existing_workflow(self):
        req = self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.post(
            reverse("administration:exceptional_absence_action", args=[req.id, "approve"]),
            {"next": reverse("administration:exceptional_absences")},
        )
        self.assertEqual(response.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.approval_stage, StaffRequest.APPROVAL_DIRECTION)
        self.assertTrue(req.is_exceptional_absence)

    def test_reject_uses_existing_workflow(self):
        req = self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.post(
            reverse("administration:exceptional_absence_action", args=[req.id, "reject"]),
            {"next": reverse("administration:exceptional_absences")},
        )
        self.assertEqual(response.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(req.status, StaffRequest.STATUS_REJECTED)

    def test_approve_redirects_back_to_exceptional_page(self):
        req = self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.post(
            reverse("administration:exceptional_absence_action", args=[req.id, "approve"]),
            {"next": reverse("administration:exceptional_absences")},
        )
        self.assertRedirects(
            response, reverse("administration:exceptional_absences"), status_code=302
        )

    def test_non_exceptional_requests_not_listed(self):
        ensure_leave_window(self.employee.profile)
        AnnualLeave.objects.filter(employee=self.employee.profile).update(consumed=Decimal("0"))
        self.client.logout()
        self.client.login(username="agent", password="TestPass123!")
        self.client.post(
            reverse("requests_management:leave_create"),
            {
                "start_date": "2026-07-13",
                "end_date": "2026-07-15",
                "total_days": "3",
                "remaining_days_for_reason": "",
                "reason": "Conge normal",
            },
            follow=True,
        )
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aucune absence exceptionnelle trouvee.")

    def test_balance_not_modified_after_approval(self):
        req = self._create_exceptional_request(days="3")
        initial_balance = self.employee.profile.leave_balance
        self._get_admin()
        self.client.post(
            reverse("administration:exceptional_absence_action", args=[req.id, "approve"]),
            {"next": reverse("administration:exceptional_absences")},
        )
        self.employee.profile.refresh_from_db()
        self.assertEqual(self.employee.profile.leave_balance, initial_balance)

    def test_summary_stats_displayed(self):
        self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total absences exceptionnelles")
        self.assertContains(response, "En attente")
        self.assertContains(response, "Approuvees")
        self.assertContains(response, "Jours de retenue")

    def test_button_present_in_presence_page(self):
        self._get_admin()
        response = self.client.get(reverse("administration:presence_overview"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Absences exceptionnelles")

    def test_nav_link_present_when_on_page(self):
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Absences exceptionnelles")

    def test_deduction_filter_pending(self):
        """Filter by pending salary deduction status."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.save(update_fields=["approval_stage", "status"])
        self._get_admin()
        response = self.client.get(
            reverse("administration:exceptional_absences") + "?deduction=pending"
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agent")

    def test_deduction_filter_withdrawn(self):
        """Filter by withdrawn salary deduction status."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.salary_deduction_status = StaffRequest.DEDUCTION_STATUS_WITHDRAWN
        req.save(update_fields=["approval_stage", "status", "salary_deduction_status"])
        self._get_admin()
        response = self.client.get(
            reverse("administration:exceptional_absences") + "?deduction=withdrawn"
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Agent")

    def test_mark_salary_deduction_withdrawn(self):
        """Admin can mark a pending salary deduction as withdrawn."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.save(update_fields=["approval_stage", "status"])
        initial_status = req.salary_deduction_status
        self._get_admin()
        response = self.client.post(
            reverse("administration:mark_salary_deduction", args=[req.id])
        )
        self.assertEqual(response.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(
            req.salary_deduction_status, StaffRequest.DEDUCTION_STATUS_WITHDRAWN
        )
        self.assertNotEqual(req.salary_deduction_status, initial_status)
        history = RequestActionHistory.objects.filter(request=req).last()
        self.assertIsNotNone(history)

    def test_mark_salary_deduction_requires_admin(self):
        """Non-admin cannot mark salary deduction."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.save(update_fields=["approval_stage", "status"])
        self.client.logout()
        self.client.login(username="regular_agent", password="TestPass123!")
        response = self.client.post(
            reverse("administration:mark_salary_deduction", args=[req.id])
        )
        self.assertEqual(response.status_code, 302)

    def test_mark_salary_deduction_not_allowed_for_unapproved(self):
        """Cannot mark salary deduction for an unapproved request."""
        req = self._create_exceptional_request(days="3")
        self._get_admin()
        response = self.client.post(
            reverse("administration:mark_salary_deduction", args=[req.id])
        )
        self.assertEqual(response.status_code, 302)
        req.refresh_from_db()
        self.assertEqual(
            req.salary_deduction_status, StaffRequest.DEDUCTION_STATUS_PENDING
        )

    def test_print_view(self):
        """Print view returns A4 printable page."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.save(update_fields=["approval_stage", "status"])
        self._get_admin()
        response = self.client.get(
            reverse("administration:exceptional_absence_print", args=[req.id])
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impression")
        self.assertContains(response, "Mamy Agent")

    def test_deduction_stats_in_data_view(self):
        """Data view returns deduction counts in JSON."""
        req = self._create_exceptional_request(days="3")
        req.approval_stage = StaffRequest.APPROVAL_DIRECTION
        req.status = StaffRequest.STATUS_APPROVED
        req.save(update_fields=["approval_stage", "status"])
        self._get_admin()
        response = self.client.get(reverse("administration:exceptional_absences_data"))
        self.assertEqual(response.status_code, 200)
        if hasattr(response, "json"):
            data = response.json()
        else:
            data = json.loads(response.content)
        self.assertIn("deduction_pending_count", data)
        self.assertEqual(data["deduction_pending_count"], 1)
        self.assertEqual(data["deduction_withdrawn_count"], 0)


class LogsManagementTests(TestCase):
    """Tests for the RH logs console interface."""

    def setUp(self):
        from django.utils import timezone
        from apps.administration.models import LogEntry

        admin_role = get_role_by_code(EmployeeProfile.ROLE_ADMIN)
        user_role = get_role_by_code(EmployeeProfile.ROLE_USER)

        self.admin = User.objects.create_user(username="rh", password="TestPass123!")
        self.admin.is_staff = True
        self.admin.save()
        self.admin.profile.role = admin_role
        self.admin.profile.save()
        self.client.login(username="rh", password="TestPass123!")

        self.regular_user = User.objects.create_user(
            username="agent", password="TestPass123!", first_name="Test", last_name="User"
        )
        self.regular_user.profile.role = user_role
        self.regular_user.profile.save()

        self.log1 = LogEntry.objects.create(
            level=LogEntry.LEVEL_INFO,
            message="User logged in",
            user=self.admin,
            user_display="admin",
            request_path="/admin-metier/tableau-de-bord/",
            request_method="GET",
            request_ip="192.168.1.1",
            logger_name="application",
        )
        self.log2 = LogEntry.objects.create(
            level=LogEntry.LEVEL_ERROR,
            message="Database error occurred",
            user=self.regular_user,
            user_display="agent",
            request_path="/admin-metier/absences-exceptionnelles/",
            request_method="GET",
            request_ip="192.168.1.2",
            logger_name="application",
            error_id="ERR-20260903-ABCD12",
            exception_type="OperationalError",
            traceback_data="Traceback (most recent call last):\n  File ...",
        )
        self.log3 = LogEntry.objects.create(
            level=LogEntry.LEVEL_WARNING,
            message="Configuration warning",
            user=None,
            user_display="",
            request_path="/admin-metier/parametres/",
            request_method="POST",
            request_ip="",
            logger_name="application.notifications",
        )
        now = timezone.now()
        LogEntry.objects.filter(id=self.log1.id).update(created_at=now - timezone.timedelta(hours=1))
        LogEntry.objects.filter(id=self.log2.id).update(created_at=now - timezone.timedelta(hours=2))
        LogEntry.objects.filter(id=self.log3.id).update(created_at=now - timezone.timedelta(hours=3))

    def test_logs_page_accessible_to_admin(self):
        """Admin can access logs page."""
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 200)

    def test_logs_page_not_accessible_to_regular_user(self):
        """Regular user is redirected from logs page."""
        self.client.logout()
        self.client.login(username="agent", password="TestPass123!")
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 302)

    def test_logs_page_not_accessible_when_unauthenticated(self):
        """Unauthenticated user is redirected to login."""
        self.client.logout()
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 302)

    def test_logs_summary_stats_displayed(self):
        """Summary stats are displayed on logs page."""
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total logs")
        self.assertContains(response, "INFO")
        self.assertContains(response, "WARNING")
        self.assertContains(response, "ERROR")

    def test_logs_search_by_message(self):
        """Search finds logs by message content."""
        response = self.client.get(reverse("administration:logs") + "?q=Database")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Database error occurred")

    def test_logs_search_by_user(self):
        """Search finds logs by username."""
        response = self.client.get(reverse("administration:logs") + "?q=admin")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "User logged in")

    def test_logs_search_by_url(self):
        """Search finds logs by URL."""
        response = self.client.get(reverse("administration:logs") + "?q=absences")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Database error occurred")

    def test_logs_filter_by_level(self):
        """Filter logs by level."""
        response = self.client.get(reverse("administration:logs") + "?level=error")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Database error occurred")
        self.assertNotContains(response, "User logged in")

    def test_logs_filter_by_user(self):
        """Filter logs by user."""
        response = self.client.get(reverse("administration:logs") + f"?user={self.regular_user.id}")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Database error occurred")
        self.assertNotContains(response, "User logged in")

    def test_logs_filter_by_method(self):
        """Filter logs by HTTP method."""
        response = self.client.get(reverse("administration:logs") + "?method=POST")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Configuration warning")

    def test_logs_combined_filters(self):
        """Combine level and method filters."""
        response = self.client.get(reverse("administration:logs") + "?level=warning&method=POST")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Configuration warning")

    def test_logs_sort_by_date_descending(self):
        """Default sort is newest first."""
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        # Check that log1 (most recent) appears before log2 and log3 in the table
        row1_start = content.find(f'href="{reverse("administration:log_detail", args=[self.log1.id])}"')
        row2_start = content.find(f'href="{reverse("administration:log_detail", args=[self.log2.id])}"')
        row3_start = content.find(f'href="{reverse("administration:log_detail", args=[self.log3.id])}"')
        self.assertGreaterEqual(row1_start, 0)
        self.assertGreaterEqual(row2_start, 0)
        self.assertGreaterEqual(row3_start, 0)
        self.assertLess(row1_start, row2_start)
        self.assertLess(row2_start, row3_start)

    def test_logs_pagination(self):
        """Pagination works on logs page."""
        from django.utils import timezone
        from apps.administration.models import LogEntry

        for i in range(35):
            LogEntry.objects.create(
                level=LogEntry.LEVEL_INFO,
                message=f"Test log {i}",
                user_display="test",
                request_path="/test/",
                request_method="GET",
                created_at=timezone.now() - timezone.timedelta(minutes=i),
            )
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Suivant")

    def test_log_detail_view(self):
        """Log detail shows complete information."""
        response = self.client.get(reverse("administration:log_detail", args=[self.log2.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "ERR-20260903-ABCD12")
        self.assertContains(response, "Database error occurred")
        self.assertContains(response, "Traceback")
        self.assertContains(response, "OperationalError")
        self.assertContains(response, "Copier")

    def test_log_detail_view_permission_denied_for_regular_user(self):
        """Regular user cannot view log detail."""
        self.client.logout()
        self.client.login(username="agent", password="TestPass123!")
        response = self.client.get(reverse("administration:log_detail", args=[self.log1.id]))
        self.assertEqual(response.status_code, 302)

    def test_logs_data_view(self):
        """Ajax data view returns JSON with summary and rows."""
        response = self.client.get(reverse("administration:logs_data"))
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertIn("summary", data)
        self.assertIn("rows_html", data)
        self.assertEqual(data["summary"]["total_count"], 3)

    def test_logs_no_passwords_in_display(self):
        """Passwords/tokens are never exposed in logs."""
        from apps.administration.models import LogEntry

        sensitive_log = LogEntry.objects.create(
            level=LogEntry.LEVEL_ERROR,
            message="password=secret123 token=abc456 cookie=sessionid=xyz",
            user_display="test",
            request_path="/test/",
            request_method="GET",
        )
        response = self.client.get(reverse("administration:log_detail", args=[sensitive_log.id]))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "password=secret123")
        self.assertNotContains(response, "token=abc456")

    def test_logs_data_view_paginates_with_ajax(self):
        """Ajax data view with ajax=1 param still works for polling."""
        response = self.client.get(reverse("administration:logs_data") + "?ajax=1")
        self.assertEqual(response.status_code, 200)

    def test_logs_clear_all_requires_permission(self):
        """Non-admin cannot clear all logs."""
        self.client.logout()
        self.client.login(username="agent", password="TestPass123!")
        response = self.client.post(reverse("administration:logs_clear"), {"strategy": "all"})
        self.assertEqual(response.status_code, 302)

    def test_logs_clear_all_as_admin(self):
        """Admin can clear all logs."""
        from apps.administration.models import LogEntry

        total = LogEntry.objects.count()
        response = self.client.post(reverse("administration:logs_clear"), {"strategy": "all"})
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertTrue(data["success"])
        self.assertEqual(data["deleted"], total)
        self.assertEqual(LogEntry.objects.count(), 0)

    def test_logs_clear_all_preserves_audit_trail(self):
        """Clearing all logs preserves the audit entry in AccountActionHistory."""
        from apps.administration.models import AccountActionHistory

        self.client.post(reverse("administration:logs_clear"), {"strategy": "all"})
        audit_exists = AccountActionHistory.objects.filter(
            action=AccountActionHistory.ACTION_UPDATED,
        ).exists()
        self.assertTrue(audit_exists)

    def test_logs_clear_older_than_as_admin(self):
        """Admin can clear logs older than X days."""
        from apps.administration.models import LogEntry
        from django.utils import timezone

        LogEntry.objects.create(
            level=LogEntry.LEVEL_INFO,
            message="Old log for testing",
            user_display="test",
            request_path="/test/",
            request_method="GET",
        )
        now = timezone.now()
        LogEntry.objects.filter(message="Old log for testing").update(
            created_at=now - timezone.timedelta(days=400)
        )
        response = self.client.post(
            reverse("administration:logs_clear"),
            {"strategy": "older_than", "days": "90"},
        )
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.content)
        self.assertTrue(data["success"])
        self.assertGreater(data["deleted"], 0)

    def test_logs_clear_invalid_strategy(self):
        """Invalid strategy returns error."""
        self.client.logout()
        self.client.login(username="rh", password="TestPass123!")
        response = self.client.post(
            reverse("administration:logs_clear"),
            {"strategy": "invalid"},
        )
        self.assertEqual(response.status_code, 400)

    def test_logs_clear_requires_post(self):
        """GET request to clear logs is not allowed."""
        response = self.client.get(reverse("administration:logs_clear"))
        self.assertEqual(response.status_code, 405)

    def test_logs_page_shows_clear_button(self):
        """Logs page shows the clear button for admin."""
        response = self.client.get(reverse("administration:logs"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Effacer les logs")

    def test_logs_modal_present(self):
        """Logs page includes the confirmation modal."""
        response = self.client.get(reverse("administration:logs"))
        self.assertContains(response, "logsClearModal")
        self.assertContains(response, "Supprimer les logs ?")

    def test_logs_auto_refresh_endpoint(self):
        """Logs view supports AJAX polling via ?ajax=1."""
        response = self.client.get(reverse("administration:logs") + "?ajax=1")
        self.assertEqual(response.status_code, 200)

    def test_password_masked_in_list_view(self):
        """Sensitive data is masked in the logs list table."""
        response = self.client.get(reverse("administration:logs"))
        self.assertNotContains(response, "password=secret123")

    def test_exception_type_in_extra_data(self):
        """Log detail displays exception type for error logs."""
        response = self.client.get(reverse("administration:log_detail", args=[self.log2.id]))
        self.assertContains(response, "OperationalError")

    def test_log_filter_by_level_critical(self):
        """Filter logs by CRITICAL level."""
        from apps.administration.models import LogEntry

        LogEntry.objects.create(
            level=LogEntry.LEVEL_CRITICAL,
            message="Critical system failure",
            user_display="system",
            request_path="/admin-metier/",
            request_method="GET",
        )
        response = self.client.get(reverse("administration:logs") + "?level=critical")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Critical system failure")

